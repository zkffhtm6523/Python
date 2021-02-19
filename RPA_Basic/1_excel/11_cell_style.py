from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Side
from openpyxl.styles.colors import Color
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1행의 높이를 50을 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
# 글자 색은 빨갛게, 이탤릭, 두껍게
a1.font = Font(Color="FF0000", italic=True, bold=True)
# Strike = 취소선, Font는 Arial
b1.font = Font(Color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1: # A 번호열은 제외
            continue
        # cell이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            # 배경을 초록색으로
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            # 셀의 폰트 설정
            cell.font = Font(color="FF0000")

# 틀 고정
# B2 기준으로 틀 고정
ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")