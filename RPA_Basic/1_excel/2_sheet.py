from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" #Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB 형태로 값을 변경

#생성하면서 시트 이름 바로 변경
ws1 = wb.create_sheet("YourSheet")
#2번째 index에 sheet를 생성
ws2 = wb.create_sheet("NewSheet", 2)

#Dict 형태로 Sheet에 접근
new_ws = wb["NewSheet"]

print(wb.sheetnames) # 모든 Sheet 이름 확인

#Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")