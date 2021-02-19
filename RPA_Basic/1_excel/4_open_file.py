from openpyxl import load_workbook #파일 불러오기
wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb 불러오기
ws = wb.active #활성화된 sheet

#cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4
    print()

print()
## -> 몇 번째 줄까지 있는 지 사전에 알고 있었음.

#cell 데이터 정보를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4
    print()
    