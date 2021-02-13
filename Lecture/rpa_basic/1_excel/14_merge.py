from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
# B2 부터 D2까지 합치겠음
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")