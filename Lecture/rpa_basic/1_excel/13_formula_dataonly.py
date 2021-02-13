from openpyxl import load_workbook

# 1. 수식 그대로 가져오고 있음
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)

# 2. 위쪽과 결과는 다름, 수식이 반영되지 않음
# evaluate 되지 않은 상태의 데이터는 None으로 표시
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

# 엑셀 파일을 열어서 저장하고 다시 실행하면 정상적으로 출력됨