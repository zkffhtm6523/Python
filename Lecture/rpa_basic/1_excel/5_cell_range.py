from openpyxl import Workbook
from random import randint

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호","영어","수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_b = ws["B"] # 영어 column만 가지고 오기
print(col_b)

for cell in col_b:
    print(cell.value)

col_range = ws["B:C"] # 영어, 수학 컬럼 같이 가져오기

for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range2 = ws[2:ws.max_row]
for rows in row_range2:
    for cell in rows:
        # print(cell.value, end=" ")
        print(cell.coordinate, end= " ") # A10, AZ/250
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

# 전체 rows
# print(tuple(ws.rows))
for rows in tuple(ws.rows):
    print(rows[1].value)

# 전체 columns
# print(tuple(ws.columns))

# 전체 rows
# for row in ws.iter_rows(): # 전체 rows
#     print(row)

for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): # 전체 rows
    print(row)

# 전체 columns
# for column in ws.iter_columns(): # 전체 rows
#     print(column)


wb.save("sample.xlsx")